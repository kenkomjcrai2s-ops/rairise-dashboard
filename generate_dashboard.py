#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
らいらいず 個人成績・段位進捗ダッシュボード生成スクリプト（複数ペア対応版）

使い方:
    python3 generate_dashboard.py <入力xlsxパス> <出力htmlパス>

新xlsx仕様（複数ペア対応版）:
- 「一覧」シート: 5行目がヘッダー、6行目以降がデータ。列は
  名前キー(A) / 表示名(B) / はなまるNO(C) / 審査段位(D) / 平均着順(E) /
  ラス回避率(F) / トップ回数(G) / 実践回数(H) / ポイント(I)
- 各メンバーの個人タブ（シート名 = 一覧のA列キーと一致）
  - 各「審査ペア」は連続する2行で構成される（上=素点、下=着順）
  - 1組目: 4-5行目
  - 2組目: 6-7行目
  - 3組目: 8-9行目
  - ...（以降 2行ずつ増えていく。人によりペア数が異なる）
  - 各ペアの F列（列6）に段位ラベルがあることが多い（無い場合もある）
  - 各ペアの J〜AM列（列10〜39）が 実績1〜30（該当ペアの各半荘の得点/着順）
  - BY5:CB18（列77〜80、行5〜18）に段位別の昇段基準テーブル
- 「原本」シートは無視する。

計算ルール（★ユーザー指定）:
- 累計成績 = 該当タブ内の【全ペア】の全対局データを合算した集計
- 期間内成績 = 該当タブ内の【一番下（最下段）のデータ入りペア】の集計
- 段位判定:
  - 最下段ペアの段位ラベル = 一覧の審査段位 → そのペアが現在の審査
  - 段位ラベルが空欄 → 一覧の審査段位を採用（同じ段位の継続とみなす）
  - ラベルと一覧が食い違う → その段位を突破して新段位に上がり、新段位ではまだ0半荘
    （その旨を確認事項として報告する）
"""

import sys, json, math, datetime
import pandas as pd
import openpyxl

RANK_ORDER = ["5級","4級","3級","2級","1級","初段","二段","三段","四段","五段","六段","七段","八段","九段"]


def to_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and math.isnan(v):
            return None
        return v
    return None


def extract_pairs(ws):
    """個人シートから全ペアを抽出。データが1件でもあるペアのみ返す。"""
    pairs = []
    consecutive_empty = 0
    for pair_idx in range(60):  # 十分な数の枠を確認
        r_score = 4 + pair_idx * 2
        r_place = r_score + 1
        rank_label = ws.cell(row=r_score, column=6).value
        rank_label = str(rank_label).strip() if rank_label else None
        games = []
        for c in range(10, 40):
            s = to_num(ws.cell(row=r_score, column=c).value)
            p = to_num(ws.cell(row=r_place, column=c).value)
            if s is not None and p in (1, 2, 3, 4):
                games.append({"score": s, "place": int(p)})
        if not games:
            if not rank_label:
                consecutive_empty += 1
                if consecutive_empty >= 3:
                    break  # 3枠連続で完全に空なら終端とみなす
            else:
                consecutive_empty = 0
            continue
        consecutive_empty = 0
        pairs.append({"rank": rank_label, "games": games})
    return pairs


def compute_window_stats(games):
    played = len(games)
    if played == 0:
        return {"played": 0, "avg_place": None, "avg_score": 0.0, "place_sum": 0,
                "last_avoid_rate": None, "total_score": 0}
    place_sum = sum(g["place"] for g in games)
    total_score = sum(g["score"] for g in games)
    avoid = sum(1 for g in games if g["place"] != 4)
    return {
        "played": played,
        "avg_place": round(place_sum / played, 3),
        "avg_score": round(total_score / played, 2),
        "place_sum": place_sum,
        "last_avoid_rate": round(avoid / played, 3),
        "total_score": total_score,
    }


def extract(xlsx_path):
    df = pd.read_excel(xlsx_path, sheet_name="一覧", header=None)
    raw_rows = df.iloc[5:].values.tolist()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # 段位基準テーブル（どのシートでも共通のはずなので、"一覧"・"原本"以外の最初のシートから読む）
    ref_sheet_name = next(s for s in wb.sheetnames if s not in ("一覧", "原本"))
    ref_ws = wb[ref_sheet_name]
    rank_req = {}
    for i, rank in enumerate(RANK_ORDER):
        r = 5 + i
        rank_req[rank] = {
            "games": ref_ws.cell(row=r, column=78).value,
            "place": ref_ws.cell(row=r, column=79).value,
            "avg": ref_ws.cell(row=r, column=80).value,
        }

    players = []
    seen = set()
    warnings = []

    for r in raw_rows:
        key, disp_name, hanamaru_no, rank, avg, lastavoid, top, games, pts = r
        if pd.isna(key) or str(key).strip() == '':
            continue
        key = str(key).strip()
        disp_name = None if pd.isna(disp_name) else str(disp_name).strip()
        name = disp_name if (disp_name and disp_name != "0") else key

        rank = str(rank).strip() if not pd.isna(rank) else ""
        top_n = int(top) if not pd.isna(top) else 0

        dedup_key = (key, rank)
        if dedup_key in seen:
            continue
        seen.add(dedup_key)

        exam_target = rank if rank in RANK_ORDER else None

        if key not in wb.sheetnames:
            warnings.append(f"一覧に「{key}」がありますが、同名の個人タブが見つかりません。")
            pairs = []
        else:
            pairs = extract_pairs(wb[key])

        # ---- 累計成績: 全ペアを合算 ----
        all_games = [g for p in pairs for g in p["games"]]
        career = compute_window_stats(all_games)
        career_top = sum(1 for g in all_games if g["place"] == 1)
        career["top_count"] = career_top

        # ---- 期間内成績: 最下段ペア ----
        exam = None
        if exam_target:
            if not pairs:
                exam = _make_exam(exam_target, rank_req[exam_target], [])
            else:
                last = pairs[-1]
                last_rank = last["rank"]
                if last_rank is None or last_rank == exam_target:
                    # 継続 or ラベル省略 → 最下段ペアをそのまま期間内として採用
                    used_rank = exam_target
                    exam = _make_exam(used_rank, rank_req[used_rank], last["games"])
                else:
                    # 最下段ペアの段位ラベルが一覧と食い違う → 新段位に上がったばかりで0半荘
                    warnings.append(
                        f"「{name}」: 最下段ペアのラベル({last_rank})と一覧の審査段位({exam_target})が"
                        f"食い違うため、新段位0半荘スタートとして計算しました。"
                    )
                    exam = _make_exam(exam_target, rank_req[exam_target], [])
        else:
            exam = {"rank": None, "status": "unranked"}

        players.append({
            "key": key,
            "name": name,
            "rank": rank,
            "rank_value": RANK_ORDER.index(rank) if rank in RANK_ORDER else -1,
            "career": {
                "games": career["played"],
                "avg_place": career["avg_place"],
                "last_avoid_rate": career["last_avoid_rate"],
                "points": round(career["total_score"], 1),
                "top_count": career_top,
            },
            "exam": exam,
        })

    return {"rank_order": RANK_ORDER, "rank_req": rank_req, "players": players}, warnings


def _make_exam(rank, req, games):
    stats = compute_window_stats(games)
    played = stats["played"]
    required_games = req["games"]
    required_place = req["place"]
    required_avg = req["avg"]
    remaining_games = max(required_games - played, 0)
    place_clear = (stats["avg_place"] is not None) and (stats["avg_place"] <= required_place)
    avg_clear = played > 0 and stats["avg_score"] >= required_avg
    total_required = required_avg * required_games
    total_remaining = max(total_required - stats["total_score"], 0)
    pace_needed = round(total_remaining / remaining_games, 1) if remaining_games > 0 else None

    if played == 0:
        status = "not_started"
    elif played >= required_games:
        status = "ready" if (place_clear and avg_clear) else "reset_pending"
    else:
        status = "on_track" if (place_clear and avg_clear) else "behind"

    return {
        "rank": rank,
        "rank_value": RANK_ORDER.index(rank),
        "required_games": required_games,
        "required_place": required_place,
        "required_avg": required_avg,
        "played_games": played,
        "avg_place": stats["avg_place"],
        "avg_score": stats["avg_score"],
        "place_sum": stats["place_sum"],
        "last_avoid_rate": stats["last_avoid_rate"],
        "remaining_games": remaining_games,
        "place_clear": place_clear,
        "avg_clear": avg_clear,
        "total_current": round(stats["total_score"], 1),
        "total_required": round(total_required, 1),
        "total_remaining": round(total_remaining, 1),
        "pace_needed": pace_needed,
        "status": status,
    }


TEMPLATE = r"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>らいらいず 個人成績・段位進捗</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Shippori+Mincho:wght@500;700;800&family=Zen+Kaku+Gothic+New:wght@400;500;700;900&display=swap" rel="stylesheet">
<style>
  :root{
    --felt-1:#0b3d2e; --felt-2:#0e2f26; --felt-edge:#062018;
    --tile:#f6efdd; --tile-edge:#cdbb8e;
    --ink:#20281f; --ink-soft:#5b6357;
    --red:#a8322c; --green-dragon:#1f6b3a; --gold:#b8860b;
    --amber:#b3781f;
    --shadow: 0 6px 16px rgba(0,0,0,.35);
  }
  *{box-sizing:border-box;}
  html,body{margin:0;padding:0;}
  body{
    font-family:'Zen Kaku Gothic New', sans-serif;
    color:var(--ink);
    background:
      radial-gradient(circle at 15% 10%, rgba(255,255,255,.03), transparent 40%),
      repeating-linear-gradient(45deg, rgba(255,255,255,0.015) 0 2px, transparent 2px 10px),
      linear-gradient(160deg, var(--felt-1), var(--felt-2) 60%, var(--felt-edge));
    min-height:100vh;
    padding-bottom:80px;
  }
  .wrap{max-width:1040px;margin:0 auto;padding:36px 20px 0;}

  header.plaque{text-align:center;padding:38px 20px 30px;}
  .kanban{
    display:inline-block; background:var(--tile); border:1px solid var(--tile-edge);
    border-radius:10px; box-shadow: var(--shadow), inset 0 0 0 3px rgba(255,255,255,.4);
    padding:22px 48px; position:relative;
  }
  .kanban::before, .kanban::after{content:"";position:absolute;top:8px;bottom:8px;width:1px;background:rgba(0,0,0,.08);}
  .kanban::before{left:10px;} .kanban::after{right:10px;}
  .kanban h1{font-family:'Shippori Mincho', serif;font-weight:800;font-size:2.6rem;letter-spacing:.14em;margin:0;color:var(--ink);}
  .kanban .sub{margin-top:6px;font-size:.82rem;letter-spacing:.2em;color:var(--ink-soft);}
  .updated{margin-top:14px;font-size:.78rem;color:rgba(255,255,255,.55);letter-spacing:.05em;}

  .statbar{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin:0 0 34px;}
  .stat{background:rgba(255,255,255,.06);border:1px solid rgba(255,255,255,.12);border-radius:10px;padding:14px 10px;text-align:center;color:#eef3ee;}
  .stat .n{font-family:'Shippori Mincho', serif;font-weight:700;font-size:1.5rem;color:#f6efdd;}
  .stat .l{font-size:.66rem;letter-spacing:.1em;color:rgba(255,255,255,.55);margin-top:2px;}

  h2.section-title{
    font-family:'Shippori Mincho', serif; color:#f6efdd; font-size:1.05rem;
    letter-spacing:.2em; text-align:center; margin:0 0 18px; opacity:.85;
  }
  h2.section-title .asof{
    display:block; font-family:'Zen Kaku Gothic New',sans-serif; font-weight:400;
    letter-spacing:.05em; font-size:.68rem; color:rgba(255,255,255,.5);
    margin-top:6px;
  }

  .picker{display:flex; gap:10px; flex-wrap:wrap; align-items:center; justify-content:center; margin-bottom:20px;}
  .picker select{
    background:var(--tile); border:1px solid var(--tile-edge); color:var(--ink);
    font-family:'Zen Kaku Gothic New',sans-serif; font-weight:700; font-size:.9rem;
    padding:9px 16px; border-radius:20px; cursor:pointer; min-width:220px;
  }

  .examcard{
    background:var(--tile); border:1px solid var(--tile-edge); border-radius:16px;
    box-shadow:var(--shadow); padding:26px 26px 22px; margin-bottom:46px;
  }
  .exam-head{display:flex; align-items:center; justify-content:space-between; flex-wrap:wrap; gap:10px; margin-bottom:14px;}
  .exam-name{font-family:'Shippori Mincho', serif; font-weight:800; font-size:1.5rem;}
  .exam-badges{display:flex; gap:8px; flex-wrap:wrap;}
  .rankbadge{display:inline-block;font-size:.72rem;padding:4px 12px;border-radius:20px;color:#fff;font-weight:700;letter-spacing:.05em;}
  .statusbadge{display:inline-block;font-size:.72rem;padding:4px 12px;border-radius:20px;font-weight:700;letter-spacing:.05em;border:1.5px solid;}
  .statusbadge.ready{color:var(--green-dragon);border-color:var(--green-dragon);background:#e9f5ec;}
  .statusbadge.on_track{color:#1f6b3a;border-color:#1f6b3a;background:#eef7f0;}
  .statusbadge.behind{color:var(--amber);border-color:var(--amber);background:#fbf1e3;}
  .statusbadge.reset_pending{color:var(--red);border-color:var(--red);background:#fbeceb;}
  .statusbadge.not_started{color:var(--ink-soft);border-color:var(--ink-soft);background:#f1efe8;}
  .statusbadge.unranked{color:var(--ink-soft);border-color:var(--ink-soft);background:#f1efe8;}

  .exam-msg{font-size:.82rem; color:var(--ink-soft); margin-bottom:18px; line-height:1.6;}

  .tilewall{display:flex; flex-wrap:wrap; gap:4px; margin-bottom:20px;}
  .tilewall .t{width:18px;height:24px;border-radius:3px;background:#efe6cd;border:1px solid var(--tile-edge);}
  .tilewall .t.done{background:var(--green-dragon);border-color:var(--green-dragon);}
  .tilewall .lbl{width:100%;font-size:.68rem;color:var(--ink-soft);margin-bottom:2px;}

  .metrics{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:16px;}
  .metric{background:rgba(0,0,0,.03);border:1px solid var(--tile-edge);border-radius:10px;padding:14px 16px;}
  .metric .mlabel{font-size:.7rem;color:var(--ink-soft);letter-spacing:.05em;margin-bottom:6px;display:flex;justify-content:space-between;}
  .metric .mval{font-family:'Shippori Mincho', serif;font-size:1.3rem;font-weight:700;}
  .metric .mval .req{font-size:.78rem;font-weight:500;color:var(--ink-soft);margin-left:6px;}
  .mtag{font-size:.66rem;padding:2px 8px;border-radius:12px;font-weight:700;}
  .mtag.clear{background:#e9f5ec;color:var(--green-dragon);}
  .mtag.notclear{background:#fbeceb;color:var(--red);}

  .pacebox{background:rgba(184,134,11,.08); border:1px dashed var(--gold); border-radius:10px; padding:12px 16px; font-size:.8rem; color:#6b5410; line-height:1.6;}
  .pacebox.placebox{background:rgba(31,107,58,.07); border:1px dashed var(--green-dragon); color:#204d2f; margin-top:8px;}

  .simulator{margin-top:16px; padding:16px 18px; border-radius:12px; border:2px solid var(--tile-edge); background:rgba(0,0,0,.03); transition: background .2s ease, border-color .2s ease;}
  .simulator.safe{background:rgba(31,107,58,.12); border-color:var(--green-dragon);}
  .simulator.unsafe{background:rgba(168,50,44,.10); border-color:var(--red);}
  .simulator .simtitle{font-size:.76rem; letter-spacing:.05em; color:var(--ink-soft); margin-bottom:10px; font-weight:700;}
  .simulator .siminputs{display:grid; grid-template-columns:repeat(4,1fr); gap:10px; margin-bottom:12px;}
  .simulator .simcol{text-align:center;}
  .simulator .simcol label{display:block; font-size:.7rem; color:var(--ink-soft); margin-bottom:4px;}
  .simulator .simcol input{width:100%; text-align:center; font-family:'Shippori Mincho', serif; font-weight:700; font-size:1.1rem; padding:8px 4px; border-radius:8px; border:1px solid var(--tile-edge); background:var(--tile); color:var(--ink);}
  .simulator .simresult{font-size:.82rem; line-height:1.6; color:var(--ink);}
  .simulator .simresult b{font-family:'Shippori Mincho', serif; font-size:1.05rem;}
  .simulator .simreset{margin-top:8px; background:none; border:1px solid var(--ink-soft); color:var(--ink-soft); font-size:.7rem; padding:5px 12px; border-radius:14px; cursor:pointer;}

  /* ---- career + period two-panel ---- */
  .stats-two{
    margin-top:18px; padding-top:16px; border-top:1px dashed var(--tile-edge);
    display:grid; grid-template-columns:1fr 1fr; gap:14px;
  }
  .stats-panel{
    background:rgba(0,0,0,.03); border:1px solid var(--tile-edge); border-radius:10px;
    padding:14px 14px 10px;
  }
  .stats-panel .ptitle{
    font-family:'Shippori Mincho', serif; font-weight:700; font-size:.85rem;
    letter-spacing:.1em; color:var(--ink); text-align:center;
    padding-bottom:8px; margin-bottom:10px; border-bottom:1px solid var(--tile-edge);
  }
  .stats-panel .ptitle.career{color:var(--gold);}
  .stats-panel .ptitle.period{color:var(--green-dragon);}
  .stats-panel .pgrid{display:grid; grid-template-columns:repeat(2,1fr); gap:8px 12px; text-align:center;}
  .stats-panel .pgrid > div{font-size:.68rem; color:var(--ink-soft);}
  .stats-panel .pgrid b{display:block; font-family:'Shippori Mincho', serif; font-size:1rem; color:var(--ink); font-weight:700; margin-bottom:1px;}

  .controls{display:flex; gap:10px; flex-wrap:wrap; align-items:center; justify-content:space-between; margin-bottom:14px;}
  .sortbtns{display:flex; gap:6px; flex-wrap:wrap;}
  .sortbtns button{background:rgba(255,255,255,.08); border:1px solid rgba(255,255,255,.18); color:#eef3ee; font-family:'Zen Kaku Gothic New',sans-serif; font-size:.74rem; padding:7px 12px; border-radius:20px; cursor:pointer; letter-spacing:.05em; transition:.15s;}
  .sortbtns button:hover{background:rgba(255,255,255,.16);}
  .sortbtns button.active{background:var(--tile); color:var(--ink); border-color:var(--tile); font-weight:700;}
  .searchbox{background:rgba(255,255,255,.08); border:1px solid rgba(255,255,255,.18); border-radius:20px; padding:7px 14px;}
  .searchbox input{background:transparent; border:none; outline:none; color:#f6efdd; font-size:.8rem; width:150px; font-family:'Zen Kaku Gothic New',sans-serif;}
  .searchbox input::placeholder{color:rgba(255,255,255,.4);}

  .board{background:rgba(0,0,0,.15); border-radius:14px; padding:8px; border:1px solid rgba(255,255,255,.08);}
  .row{display:grid; grid-template-columns:44px 1.3fr 78px 66px 60px 74px; align-items:center; gap:6px; background:var(--tile); border:1px solid var(--tile-edge); border-radius:8px; padding:11px 12px; margin-bottom:6px; font-size:.82rem; cursor:pointer; transition:transform .12s ease;}
  .row:hover{transform:translateX(2px);}
  .row.selected{outline:2px solid var(--gold);}
  .row.head{background:transparent;border:none;color:rgba(255,255,255,.5);font-size:.64rem;letter-spacing:.08em;padding:2px 12px 8px;margin-bottom:0;cursor:default;}
  .row.head:hover{transform:none;}
  .row .rk{font-family:'Shippori Mincho', serif;font-weight:800;font-size:.95rem;color:var(--ink-soft);text-align:center;}
  .row .nm{font-weight:700;color:var(--ink);overflow:hidden;text-overflow:ellipsis;white-space:nowrap;}
  .row .badge{justify-self:start;font-size:.66rem;padding:3px 9px;border-radius:20px;color:#fff;text-align:center;white-space:nowrap;}
  .row .num{text-align:right;color:var(--ink);font-variant-numeric:tabular-nums;}
  .row .st{justify-self:end;font-size:.62rem;padding:3px 8px;border-radius:12px;font-weight:700;text-align:center;white-space:nowrap;}
  .row .st.ready{background:#e9f5ec;color:var(--green-dragon);}
  .row .st.on_track{background:#eef7f0;color:#1f6b3a;}
  .row .st.behind{background:#fbf1e3;color:var(--amber);}
  .row .st.reset_pending{background:#fbeceb;color:var(--red);}
  .row .st.not_started{background:#f1efe8;color:var(--ink-soft);}
  .row .st.unranked{background:#f1efe8;color:var(--ink-soft);}

  .empty{text-align:center;color:rgba(255,255,255,.5);padding:40px;font-size:.85rem;}
  footer{text-align:center;color:rgba(255,255,255,.35);font-size:.7rem;margin-top:40px;letter-spacing:.05em;}

  @media (max-width:720px){
    .statbar{grid-template-columns:repeat(2,1fr);}
    .metrics{grid-template-columns:1fr;}
    .stats-two{grid-template-columns:1fr;}
    .row{grid-template-columns:30px 1fr 60px 60px;}
    .row .col-place, .row .col-status{display:none;}
    .row.head .col-place, .row.head .col-status{display:none;}
  }
</style>
</head>
<body>
<div class="wrap">

  <header class="plaque">
    <div class="kanban">
      <h1>らいらいず</h1>
      <div class="sub">個人 成績 ・ 段位 進捗</div>
    </div>
    <div class="updated">__GENERATED_AT__</div>
  </header>

  <div class="statbar" id="statbar"></div>

  <h2 class="section-title">昇 段 診 断<span class="asof">__DATE_ONLY__</span></h2>
  <div class="picker">
    <select id="playerSelect"></select>
  </div>
  <div class="examcard" id="examcard"></div>

  <h2 class="section-title">らいらいずメンバー一覧</h2>
  <div class="controls">
    <div class="sortbtns" id="sortbtns"></div>
    <div class="searchbox"><input id="search" type="text" placeholder="名前で検索..."></div>
  </div>
  <div class="board">
    <div class="row head">
      <div>順</div>
      <div>名前</div>
      <div>段位</div>
      <div class="num col-place">平均着順</div>
      <div class="num">半荘</div>
      <div class="col-status" style="text-align:right;">審査状況</div>
    </div>
    <div id="rows"></div>
  </div>

  <footer>らいらいず 個人成績管理表より自動生成 &mdash; 段位基準はサークル既定値に基づく</footer>
</div>

<script id="rawData" type="application/json">__DATA_JSON__</script>
<script>
const DATA = JSON.parse(document.getElementById('rawData').textContent);
const RANK_ORDER = DATA.rank_order;
const players = DATA.players;

function rankTier(rank){
  if(!rank) return '#8a6a3d';
  if(rank.includes('九段')) return '#a8322c';
  if(rank.includes('八段')||rank.includes('七段')) return '#b8860b';
  if(rank.includes('六段')||rank.includes('五段')) return '#8a6a20';
  if(rank.includes('段')) return '#1f6b3a';
  if(rank.includes('1級')||rank.includes('2級')) return '#5b6b7c';
  if(rank.includes('級')) return '#8a8a8a';
  return '#8a6a3d';
}

function fmt(v, digits){
  if(v===null || v===undefined || isNaN(v)) return '—';
  return v.toFixed(digits);
}
function fmtPts(v){
  if(v===null||v===undefined) return '—';
  const sign = v>0?'+':'';
  return sign + v.toFixed(1);
}

const STATUS_LABEL = {
  ready:'合格ライン到達', on_track:'基準クリア中', behind:'追い上げ中',
  reset_pending:'要再挑戦', not_started:'審査開始前', unranked:'未認定',
};

let selectedKey = null;
let activePlayers = players.filter(p=>p.rank_value>=0);

function renderStatbar(){
  const n = activePlayers.length;
  const totalGames = activePlayers.reduce((a,p)=>a+(p.career.games||0),0);
  const readyCount = activePlayers.filter(p=>p.exam && p.exam.status==='ready').length;
  const onTrackCount = activePlayers.filter(p=>p.exam && p.exam.status==='on_track').length;
  document.getElementById('statbar').innerHTML = `
    <div class="stat"><div class="n">${n}</div><div class="l">在籍人数</div></div>
    <div class="stat"><div class="n">${totalGames}</div><div class="l">累計半荘数（全員）</div></div>
    <div class="stat"><div class="n">${readyCount}</div><div class="l">合格ライン到達中</div></div>
    <div class="stat"><div class="n">${onTrackCount}</div><div class="l">基準クリア継続中</div></div>
  `;
}

function renderPicker(){
  const sel = document.getElementById('playerSelect');
  const sorted = [...players].sort((a,b)=> b.rank_value - a.rank_value || a.name.localeCompare(b.name,'ja'));
  sel.innerHTML = sorted.map(p=>`<option value="${p.key}">${p.name}（${p.rank}）</option>`).join('');
  sel.addEventListener('change', ()=>{ selectedKey = sel.value; renderExamCard(); highlightRow(); });
  if(!selectedKey) selectedKey = sorted[0].key;
  sel.value = selectedKey;
}

const PLACE_LABEL = {1:'トップ', 2:'二着', 3:'三着', 4:'ラス'};

function placeExampleHtml(ex){
  const remaining = ex.remaining_games;
  const placeSum = ex.avg_place!==null ? ex.place_sum : 0;
  const budget = ex.required_place * ex.required_games - placeSum;

  if(budget < remaining * 1 - 1e-9){
    return `<div class="pacebox placebox">着順基準：残り${remaining}半荘では、全対局トップを取っても平均着順の基準到達は数値上厳しい状況です。早めのペースアップが必要です。</div>`;
  }
  if(budget >= remaining * 4 - 1e-9){
    return `<div class="pacebox placebox">着順基準：残り${remaining}半荘は着順を問わず基準内に収まります（ポイント基準のみ意識すればOK）。</div>`;
  }
  let lower = Math.floor(budget / remaining);
  lower = Math.max(1, Math.min(3, lower));
  const higher = lower + 1;
  let nHigher = Math.floor(budget - remaining * lower + 1e-9);
  nHigher = Math.max(0, Math.min(remaining, nHigher));
  const nLower = remaining - nHigher;
  const counts = {1:0,2:0,3:0,4:0};
  counts[lower] += nLower;
  counts[higher] += nHigher;
  const parts = [1,2,3,4].map(k => `${PLACE_LABEL[k]}${counts[k]}回`).join('　');
  return `<div class="pacebox placebox">着順基準：残り${remaining}半荘の一例として、<b>${parts}</b> のペースならセーフ（平均着順${ex.required_place}以内を維持できます）。</div>`;
}

function simulatorHtml(ex){
  if(!ex.remaining_games || ex.remaining_games <= 0) return '';
  return `
  <div class="simulator" id="simulator">
    <div class="simtitle">着順基準シミュレーター　（残り ${ex.remaining_games} 半荘のうち、想定回数を入力）</div>
    <div class="siminputs">
      <div class="simcol"><label>トップ</label><input type="number" min="0" id="sim-1" value="0"></div>
      <div class="simcol"><label>二着</label><input type="number" min="0" id="sim-2" value="0"></div>
      <div class="simcol"><label>三着</label><input type="number" min="0" id="sim-3" value="0"></div>
      <div class="simcol"><label>ラス</label><input type="number" min="0" id="sim-4" value="0"></div>
    </div>
    <div class="simresult" id="simresult"></div>
    <button class="simreset" id="simreset">入力をリセット</button>
  </div>`;
}

function bindSimulator(ex){
  if(!ex.remaining_games || ex.remaining_games <= 0) return;
  const ids = [1,2,3,4].map(k=>document.getElementById('sim-'+k));
  const box = document.getElementById('simulator');
  const resultEl = document.getElementById('simresult');
  const cap = ex.remaining_games;

  ids.forEach(el => el.setAttribute('max', cap));

  function clampField(changedEl){
    let counts = ids.map(el => Math.max(0, parseInt(el.value)||0));
    let total = counts.reduce((a,b)=>a+b,0);
    if(total > cap){
      const idx = ids.indexOf(changedEl);
      const excess = total - cap;
      counts[idx] = Math.max(0, counts[idx] - excess);
      changedEl.value = counts[idx];
    } else {
      changedEl.value = counts[ids.indexOf(changedEl)];
    }
    const newCounts = ids.map(el => Math.max(0, parseInt(el.value)||0));
    const newTotal = newCounts.reduce((a,b)=>a+b,0);
    ids.forEach((el,i) => el.setAttribute('max', cap - (newTotal - newCounts[i])));
  }

  function update(){
    const counts = ids.map(el => Math.max(0, parseInt(el.value)||0));
    const entered = counts.reduce((a,b)=>a+b,0);
    const enteredSum = counts[0]*1 + counts[1]*2 + counts[2]*3 + counts[3]*4;
    const combinedGames = (ex.played_games||0) + entered;
    const combinedSum = (ex.place_sum||0) + enteredSum;
    const combinedAvg = combinedGames>0 ? combinedSum/combinedGames : null;
    const clear = combinedGames>0 && combinedAvg <= ex.required_place;

    box.classList.remove('safe','unsafe');
    box.classList.add(clear ? 'safe':'unsafe');

    let note = '';
    if(entered !== cap){
      note = `<br><span style="opacity:.7;">入力合計 ${entered} / 残り${cap}半荘（あと${cap-entered}半荘分入力できます）</span>`;
    } else {
      note = `<br><span style="opacity:.7;">入力合計 ${entered} / 残り${cap}半荘（上限まで入力済み）</span>`;
    }
    resultEl.innerHTML = `入力後の平均着順：<b>${combinedGames>0? combinedAvg.toFixed(3):'—'}</b>　／　基準 ${ex.required_place}以内　→　`
      + (clear ? '<b style="color:var(--green-dragon);">クリア</b>' : '<b style="color:var(--red);">未達</b>')
      + note;
  }

  ids.forEach(el => el.addEventListener('input', ()=>{ clampField(el); update(); }));
  document.getElementById('simreset').addEventListener('click', ()=>{
    ids.forEach(el => { el.value = 0; el.setAttribute('max', cap); });
    update();
  });
  update();
}

function statsTwoPanels(p, ex){
  const c = p.career || {};
  const avoidC = c.last_avoid_rate!==null && c.last_avoid_rate!==undefined ? Math.round(c.last_avoid_rate*100)+'%' : '—';
  const avoidP = ex.last_avoid_rate!==null && ex.last_avoid_rate!==undefined ? Math.round(ex.last_avoid_rate*100)+'%' : '—';
  return `
  <div class="stats-two">
    <div class="stats-panel">
      <div class="ptitle career">累計成績</div>
      <div class="pgrid">
        <div><b>${c.games||0}</b>累計半荘数</div>
        <div><b>${fmt(c.avg_place,2)}</b>累計平均着順</div>
        <div><b>${avoidC}</b>累計ラス回避率</div>
        <div><b>${fmtPts(c.points)}</b>累計ポイント</div>
      </div>
    </div>
    <div class="stats-panel">
      <div class="ptitle period">期間内成績</div>
      <div class="pgrid">
        <div><b>${ex.played_games||0}</b>期間内半荘数</div>
        <div><b>${fmt(ex.avg_place,2)}</b>期間内平均着順</div>
        <div><b>${avoidP}</b>期間内ラス回避率</div>
        <div><b>${fmtPts(ex.total_current)}</b>期間内ポイント</div>
      </div>
    </div>
  </div>`;
}

function renderExamCard(){
  const p = players.find(x=>x.key===selectedKey);
  const el = document.getElementById('examcard');
  if(!p){ el.innerHTML = '<div class="empty">打ち手を選択してください</div>'; return; }
  const ex = p.exam;

  if(!ex || ex.status==='unranked'){
    el.innerHTML = `
      <div class="exam-head">
        <div class="exam-name">${p.name}</div>
        <div class="exam-badges"><span class="statusbadge unranked">未認定</span></div>
      </div>
      <div class="exam-msg">まだ段位が設定されていません。対局を開始すると五級からの審査がスタートします。</div>
      ${statsTwoPanels(p, {played_games:0, avg_place:null, last_avoid_rate:null, total_current:0})}
    `;
    return;
  }

  const tier = rankTier(ex.rank);
  const tiles = Array.from({length: ex.required_games}, (_,i)=>
    `<div class="t ${i < ex.played_games ? 'done':''}"></div>`
  ).join('');

  let msg = '';
  if(ex.status==='not_started'){
    msg = `直近の昇段によりこの段位の判定はリセットされました。次の対局から ${ex.required_games} 半荘の審査が始まります。`;
  } else if(ex.status==='ready'){
    msg = `規定打数 ${ex.required_games} 半荘を消化し、平均着順・平均収支ともに基準を満たしています。昇段確定ラインに到達しました。`;
  } else if(ex.status==='reset_pending'){
    msg = `規定打数 ${ex.required_games} 半荘を消化しましたが、基準未達のため次の対局から打数がリセットされ再挑戦となります。`;
  } else {
    msg = `現在 ${ex.played_games} / ${ex.required_games} 半荘を消化中です。残り ${ex.remaining_games} 半荘で基準クリアを目指しましょう。`;
  }

  const placeTag = ex.played_games>0 ? (ex.place_clear ? '<span class="mtag clear">クリア中</span>' : '<span class="mtag notclear">未達</span>') : '';
  const avgTag = ex.played_games>0 ? (ex.avg_clear ? '<span class="mtag clear">クリア中</span>' : '<span class="mtag notclear">未達</span>') : '';

  let paceHtml = '';
  if(ex.remaining_games > 0){
    if(ex.total_remaining > 0){
      paceHtml = `<div class="pacebox">残り <b>${ex.remaining_games}半荘</b> で、1半荘あたり平均 <b>${fmtPts(ex.pace_needed)}</b> 以上のペースが必要です。（不足累計: ${fmtPts(-ex.total_remaining)}）</div>`;
    } else {
      paceHtml = `<div class="pacebox">必要累計ポイントは既に確保できています。残り <b>${ex.remaining_games}半荘</b> を消化すれば昇段判定です（平均着順の基準維持にご注意ください）。</div>`;
    }
    paceHtml += placeExampleHtml(ex);
  }

  el.innerHTML = `
    <div class="exam-head">
      <div class="exam-name">${p.name}</div>
      <div class="exam-badges">
        <span class="rankbadge" style="background:${tier}">${ex.rank} 昇段審査中</span>
        <span class="statusbadge ${ex.status}">${STATUS_LABEL[ex.status]}</span>
      </div>
    </div>
    <div class="exam-msg">${msg}</div>
    <div class="tilewall">
      <div class="lbl">半荘進捗　${ex.played_games} / ${ex.required_games}</div>
      ${tiles}
    </div>
    <div class="metrics">
      <div class="metric">
        <div class="mlabel"><span>平均着順（${ex.required_place}以内）</span>${placeTag}</div>
        <div class="mval">${fmt(ex.avg_place,3)}<span class="req">/ 基準 ${ex.required_place}</span></div>
      </div>
      <div class="metric">
        <div class="mlabel"><span>平均収支（${ex.required_avg}以上）</span>${avgTag}</div>
        <div class="mval">${fmt(ex.avg_score,1)}<span class="req">/ 基準 ${ex.required_avg}</span></div>
      </div>
    </div>
    ${paceHtml}
    ${simulatorHtml(ex)}
    ${statsTwoPanels(p, ex)}
  `;
  bindSimulator(ex);
}

const sortOptions = [
  {key:'rank_value', label:'段位順', dir:-1},
  {key:'points', label:'ポイント順', dir:-1},
  {key:'avg_place', label:'平均着順順', dir:1},
  {key:'games', label:'半荘数順', dir:-1},
];
let sortKey = 'rank_value';
let sortDir = -1;

function renderSortButtons(){
  const el = document.getElementById('sortbtns');
  el.innerHTML = sortOptions.map(o=>`<button data-key="${o.key}" data-dir="${o.dir}">${o.label}</button>`).join('');
  el.querySelectorAll('button').forEach(btn=>{
    if(btn.dataset.key===sortKey) btn.classList.add('active');
    btn.addEventListener('click', ()=>{
      sortKey = btn.dataset.key; sortDir = parseInt(btn.dataset.dir);
      el.querySelectorAll('button').forEach(b=>b.classList.remove('active'));
      btn.classList.add('active');
      renderRows();
    });
  });
}

function highlightRow(){
  document.querySelectorAll('.row[data-key]').forEach(r=>{
    r.classList.toggle('selected', r.dataset.key===selectedKey);
  });
}

function playerSortValue(p, key){
  if(key==='rank_value') return p.rank_value;
  if(key==='points') return p.career.points;
  if(key==='avg_place') return p.career.avg_place;
  if(key==='games') return p.career.games;
  return 0;
}

function renderRows(){
  const q = document.getElementById('search').value.trim().toLowerCase();
  let list = activePlayers.filter(p=>p.name.toLowerCase().includes(q));
  list.sort((a,b)=>{
    let av = playerSortValue(a, sortKey), bv = playerSortValue(b, sortKey);
    if(av===null) av = -Infinity;
    if(bv===null) bv = -Infinity;
    if(av===bv) return (b.career.points||0) - (a.career.points||0);
    return (av-bv) * sortDir;
  });
  const rowsEl = document.getElementById('rows');
  if(list.length===0){ rowsEl.innerHTML = '<div class="empty">該当する打ち手が見つかりません</div>'; return; }
  rowsEl.innerHTML = list.map((p,i)=>{
    const tier = rankTier(p.rank);
    const status = p.exam ? p.exam.status : 'unranked';
    return `
    <div class="row" data-key="${p.key}">
      <div class="rk">${i+1}</div>
      <div class="nm">${p.name}</div>
      <div><span class="badge" style="background:${tier}">${p.rank}</span></div>
      <div class="num col-place">${fmt(p.career.avg_place,2)}</div>
      <div class="num">${p.career.games||0}</div>
      <div class="st col-status ${status}">${STATUS_LABEL[status]}</div>
    </div>`;
  }).join('');
  rowsEl.querySelectorAll('.row').forEach(r=>{
    r.addEventListener('click', ()=>{
      selectedKey = r.dataset.key;
      document.getElementById('playerSelect').value = selectedKey;
      renderExamCard();
      highlightRow();
      document.getElementById('examcard').scrollIntoView({behavior:'smooth', block:'center'});
    });
  });
  highlightRow();
}

document.getElementById('search').addEventListener('input', renderRows);

renderStatbar();
renderPicker();
renderExamCard();
renderSortButtons();
renderRows();
</script>
</body>
</html>
"""


def build_html(data):
    data_json = json.dumps(data, ensure_ascii=False)
    generated_at = datetime.datetime.now().strftime("%Y年%m月%d日 %H:%M 更新")
    date_only = datetime.datetime.now().strftime("%Y年%m月%d日時点")
    return (TEMPLATE
            .replace("__DATA_JSON__", data_json)
            .replace("__GENERATED_AT__", generated_at)
            .replace("__DATE_ONLY__", date_only))


def main():
    if len(sys.argv) < 3:
        print("使い方: python3 generate_dashboard.py <入力xlsx> <出力html>")
        sys.exit(1)
    xlsx_path = sys.argv[1]
    out_path = sys.argv[2]
    data, warnings = extract(xlsx_path)
    html = build_html(data)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"生成完了: {out_path}（{len(data['players'])}名分）")
    if warnings:
        print("--- 確認事項 ---")
        for w in warnings:
            print("・" + w)


if __name__ == "__main__":
    main()
